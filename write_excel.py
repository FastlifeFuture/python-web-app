import openpyxl

def write_excel(date, region, site_name, site_number, type,address,city, county, state, zip_code, s_power_system_b, s_rects_conv_b, s_load_cap_b, s_load_current_b, s_dist_b, b_power_system_b, b_rects_conv_b, b_load_cap_b, b_load_current_b, b_dist_b, s_power_system_a, s_rects_conv_a, s_load_cap_a, s_load_current_a, s_dist_a, b_power_system_a, b_rects_conv_a, b_load_cap_a, b_load_current_a, b_dist_a, manufacturer_b, type_model_b, Sulf_gel_b, batteries_b, cells_b, runtime_b, manufacturer_a, type_model_a, Sulf_gel_a, batteries_a, cells_a, runtime_a):
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb['Sheet1']

    ws['B4'].value = date
    ws['B6'].value = region
    ws['B7'].value = site_name
    ws['B8'].value = site_number
    ws['B10'].value = type
    ws['B12'].value = city 
    ws['B11'].value = address
    ws['B13'].value = county
    ws['B14'].value = state
    ws['B15'].value = zip_code

    ws['B60'].value = s_power_system_b
    ws['B61'].value = s_rects_conv_b
    ws['B62'].value = s_load_cap_b
    ws['B63'].value = s_load_current_b
    ws['B64'].value = s_dist_b
    ws['B65'].value = b_power_system_b 
    ws['B66'].value = b_rects_conv_b
    ws['B67'].value = b_load_cap_b
    ws['B68'].value = b_load_current_b
    ws['B69'].value = b_dist_b
    
    ws['C60'].value = s_power_system_a
    ws['C61'].value = s_rects_conv_a
    ws['C62'].value = s_load_cap_a
    ws['C63'].value = s_load_current_a
    ws['C64'].value = s_dist_a
    ws['C65'].value = b_power_system_a 
    ws['C66'].value = b_rects_conv_a
    ws['C67'].value = b_load_cap_a
    ws['C68'].value = b_load_current_a
    ws['C69'].value = b_dist_a

    ws['B71'].value = manufacturer_b
    ws['B72'].value = type_model_b
    ws['B73'].value = Sulf_gel_b
    ws['B74'].value = batteries_b
    ws['B75'].value = cells_b
    ws['B77'].value = runtime_b

    ws['C71'].value = manufacturer_a
    ws['C72'].value = type_model_a
    ws['C73'].value = Sulf_gel_a
    ws['C74'].value = batteries_a
    ws['C75'].value = cells_a
    ws['C76'].value = runtime_a



    wb.save('test.xlsx')


